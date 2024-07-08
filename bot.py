import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
import re
from fuzzywuzzy import fuzz, process

API_TOKEN = '7422047058:AAHaXwTV9D7hsS7wMpSc4wqrxlgzvIjIrP8'  # Замените 'YOUR_BOT_API_TOKEN' на токен вашего бота

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)

# Словарь для хранения доступных баз данных
databases = {
    'Lada Samara': 'car_repair_samara.xlsx',
    'Toyota Corolla': 'car_repair_corolla.xlsx',
    'Honda Civic': 'car_repair_civic.xlsx',
    'Vaz 2107': 'car_repair_vaz_seven.xlsx'
}

# Словарь для хранения выбранных баз данных пользователей
user_databases = {}

# Путь к файлу с неотвеченными запросами
unanswered_requests_file = 'unanswered_requests.xlsx'

def normalize_text(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s]', '', text)  # Удаляем все знаки препинания
    return text

def load_data_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
        model, issue, solution = row
        data.append((model, issue, solution))
    return data

def save_data_to_excel(file_path, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Car Issues"
    ws.append(["Model", "Issue", "Solution"])  # Заголовки колонок

    for row in data:
        ws.append(row)

    wb.save(file_path)

def update_data_in_excel(file_path, model, issue, solution):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append([model, issue, solution])
    wb.save(file_path)

def delete_data_from_excel(file_path, model, issue):
    data = load_data_from_excel(file_path)
    normalized_model = normalize_text(model)
    normalized_issue = normalize_text(issue)
    new_data = [row for row in data if not (normalize_text(row[0]) == normalized_model and normalize_text(row[1]) == normalized_issue)]
    save_data_to_excel(file_path, new_data)

def get_best_match(query, choices):
    best_match = process.extractOne(query, choices, scorer=fuzz.token_sort_ratio)
    return best_match

def save_unanswered_request(user_id, query):
    try:
        wb = openpyxl.load_workbook(unanswered_requests_file)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["User ID", "Query"])

    ws.append([user_id, query])
    wb.save(unanswered_requests_file)

@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for db_name in databases.keys():
        keyboard.add(types.KeyboardButton(db_name))
    await message.reply("Привет! Выберите модель автомобиля:", reply_markup=keyboard)

@dp.message_handler(lambda message: message.text in databases.keys())
async def select_database(message: types.Message):
    user_databases[message.from_user.id] = databases[message.text]
    await message.reply(f"Вы выбрали базу данных для {message.text}. Как я могу вам помочь?", reply_markup=types.ReplyKeyboardRemove())

@dp.message_handler(commands=['add'])
async def add_issue(message: types.Message):
    user_db = user_databases.get(message.from_user.id)
    if not user_db:
        await message.reply("Сначала выберите базу данных с помощью команды /start.")
        return
    try:
        _, model, issue, solution = message.text.split(';', 3)
        update_data_in_excel(user_db, model.strip(), issue.strip(), solution.strip())
        await message.reply("Новая запись успешно добавлена.")
    except ValueError:
        await message.reply("Использование команды: /add ;Модель;Вопрос;Решение")

@dp.message_handler(commands=['delete'])
async def delete_issue(message: types.Message):
    user_db = user_databases.get(message.from_user.id)
    if not user_db:
        await message.reply("Сначала выберите базу данных с помощью команды /start.")
        return
    try:
        _, model, issue = message.text.split(';', 2)
        delete_data_from_excel(user_db, model.strip(), issue.strip())
        await message.reply("Запись успешно удалена.")
    except ValueError:
        await message.reply("Использование команды: /delete ;Модель;Вопрос")

@dp.message_handler(commands=['unanswered'])
async def send_unanswered_requests(message: types.Message):
    try:
        wb = openpyxl.load_workbook(unanswered_requests_file)
        ws = wb.active
        unanswered_requests = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            unanswered_requests.append(f"User ID: {row[0]}, Query: {row[1]}")
        
        if unanswered_requests:
            await message.reply("\n".join(unanswered_requests))
        else:
            await message.reply("Нет неотвеченных запросов.")
    except FileNotFoundError:
        await message.reply("Нет неотвеченных запросов.")

@dp.message_handler(commands=['answer'])
async def answer_unanswered_request(message: types.Message):
    try:
        _, user_id, query, answer = message.text.split(';', 3)
        update_data_in_excel('car_repair.xlsx', 'Unknown', query.strip(), answer.strip())
        await bot.send_message(chat_id=int(user_id.strip()), text=f"Ответ на ваш запрос '{query.strip()}': {answer.strip()}")
        await message.reply("Ответ добавлен и отправлен пользователю.")
    except ValueError:
        await message.reply("Использование команды: /answer ;User ID;Query;Answer")

@dp.message_handler()
async def handle_message(message: types.Message):
    user_db = user_databases.get(message.from_user.id)
    if not user_db:
        await message.reply("Сначала выберите базу данных с помощью команды /start.")
        return

    user_query = normalize_text(message.text)
    data = load_data_from_excel(user_db)
    issues = [normalize_text(row[1]) for row in data]
    best_match, score = get_best_match(user_query, issues)

    if score > 70:
        for row in data:
            if normalize_text(row[1]) == best_match:
                await message.reply(row[2])
                break
    else:
        await message.reply("Извините, я не смог найти решение для вашей проблемы. Я сохраню ваш запрос и попробую найти ответ позже.")
        save_unanswered_request(message.from_user.id, message.text)

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
