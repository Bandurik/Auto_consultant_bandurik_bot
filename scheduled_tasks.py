import openpyxl
from apscheduler.schedulers.blocking import BlockingScheduler
from aiogram import Bot
import asyncio
import requests
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Настройки Telegram
API_TOKEN = '7422047058:AAHaXwTV9D7hsS7wMpSc4wqrxlgzvIjIrP8'  # Замените 'YOUR_BOT_API_TOKEN' на токен вашего бота
CHAT_ID = '1218691055'  # Замените 'YOUR_CHAT_ID' на ID вашего чата

bot = Bot(token=API_TOKEN)

# Путь к файлу с неотвеченными запросами
unanswered_requests_file = 'unanswered_requests.xlsx'

def update_data_in_excel(file_path, model, issue, solution):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append([model, issue, solution])
    wb.save(file_path)

async def send_unanswered_requests():
    try:
        wb = openpyxl.load_workbook(unanswered_requests_file)
        ws = wb.active
        unanswered_requests = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, query = row
            unanswered_requests.append({'user_id': user_id, 'query': query})
        
        if unanswered_requests:
            for request in unanswered_requests:
                await bot.send_message(chat_id=CHAT_ID, text=f"User ID: {request['user_id']}, Query: {request['query']}")
                logger.info(f"Отправлено сообщение: User ID: {request['user_id']}, Query: {request['query']}")
                # Отправляем запрос на генерацию ответа
                response = requests.post('http://localhost:5000/generate_answer', json={'query': request['query']})
                if response.status_code == 200:
                    answer = response.json().get('answer')
                    await bot.send_message(chat_id=CHAT_ID, text=f"Answer: {answer}")
                    logger.info(f"Сгенерированный ответ: {answer}")
                    # Добавляем ответ в базу данных
                    update_data_in_excel('car_repair.xlsx', 'Unknown', request['query'], answer)
                else:
                    logger.error("Ошибка при генерации ответа")
        else:
            logger.info("Нет неотвеченных запросов.")
    except FileNotFoundError:
        logger.info("Нет неотвеченных запросов.")

def send_unanswered_requests_sync():
    asyncio.run(send_unanswered_requests())

scheduler = BlockingScheduler()
scheduler.add_job(send_unanswered_requests_sync, 'interval', minutes=1)  # Проверяем и отправляем каждые 1 минуту

if __name__ == '__main__':
    logger.info("Запуск планировщика...")
    scheduler.start()
